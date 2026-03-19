"""Parsers for audited HHSC waitlist and cost datasets."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "march": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "may": 5,
    "june": 6,
    "jun": 6,
    "july": 7,
    "jul": 7,
    "aug": 8,
    "august": 8,
    "sept": 9,
    "sep": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

PROGRAM_MAP = {
    "CLASS": "CLASS",
    "DBMD": "DBMD",
    "HCS": "HCS",
    "HCS 1": "HCS",
    "MDCP": "MDCP",
    "STAR+": "STAR_PLUS_HCBS",
    "STAR+PLUS": "STAR_PLUS_HCBS",
    "TXHML": "TXHML",
    "TXHML 1": "TXHML",
    "TOTALS": "ALL_PROGRAMS",
}

SETTING_BLOCKS = [
    (
        "State Operated ICF/IID (SSLC)a",
        "STATE_OPERATED_ICF_IID_SSLC",
        [
            "Long-term Care Costs – Average",
            "Administrative/Overhead Costs",
            "Total State Operated ICF/IID Costs",
        ],
        [
            "long_term_care_cost_average",
            "administrative_overhead_costs",
            "total_costs",
        ],
    ),
    (
        "Non-State Operated ICF/IID (Community ICF/IID)",
        "COMMUNITY_ICF_IID",
        [
            "Long-term Care Costs – Average",
            "Acute Care Cost - Average",
            "Total Non-State Operated ICF/IID Costs",
        ],
        [
            "long_term_care_cost_average",
            "acute_care_cost_average",
            "total_costs",
        ],
    ),
    (
        "HCS Waiver: Residential",
        "HCS_RESIDENTIAL",
        [
            "Long-term Care Costs – Average",
            "Acute Care Cost - Average",
            "Total HCS Waiver: Residential Costs",
        ],
        [
            "long_term_care_cost_average",
            "acute_care_cost_average",
            "total_costs",
        ],
    ),
    (
        "HCS Waiver: Non-Residential",
        "HCS_NON_RESIDENTIAL",
        [
            "Long-term Care Costs – Average",
            "Acute Care Cost - Average",
            "Total HCS Waiver: Non-Residential Costs",
        ],
        [
            "long_term_care_cost_average",
            "acute_care_cost_average",
            "total_costs",
        ],
    ),
    (
        "HCS Waiver: All Settings",
        "HCS_ALL_SETTINGS",
        [
            "Long-term Care Costs – Average",
            "Acute Care Cost - Average",
            "Total HCS Waiver: All Settings Costs",
        ],
        [
            "long_term_care_cost_average",
            "acute_care_cost_average",
            "total_costs",
        ],
    ),
    (
        "TxHmL Waiver",
        "TXHML",
        [
            "Long-term Care Costs – Average",
            "Acute Care Cost - Average",
            "Total TxHmL Costs",
        ],
        [
            "long_term_care_cost_average",
            "acute_care_cost_average",
            "total_costs",
        ],
    ),
]

LON_LABELS = {
    "Intermittent (LON 1)": 1,
    "Limited (LON 5)": 5,
    "Extensive (LON 8)": 8,
    "Pervasive (LON 6)": 6,
    "Pervasive Plus (LON 9)": 9,
    "Overall - Total": "ALL",
    "Overall - Average": "ALL",
}

HCS_COST_AREAS = [
    {
        "program": "HCS_TXHML",
        "cost_area": "RESIDENTIAL_SL_RSS",
        "source_file": "2024-hcs-cost-report-instructions.txt",
        "source_line_start": 4071,
        "source_line_end": 4083,
        "description": "Group-home-specific supervised living and residential support costs.",
    },
    {
        "program": "HCS_TXHML",
        "cost_area": "INDIVIDUALIZED_SKILLS_SOCIALIZATION",
        "source_file": "2024-hcs-cost-report-instructions.txt",
        "source_line_start": 4084,
        "source_line_end": 4102,
        "description": "Individualized Skills and Socialization facility and staffing costs.",
    },
    {
        "program": "HCS_TXHML",
        "cost_area": "PROGRAM_ADMINISTRATION_OPERATIONS",
        "source_file": "2024-hcs-cost-report-instructions.txt",
        "source_line_start": 4106,
        "source_line_end": 4112,
        "description": "Direct program-management administrative expenses for the contracted provider.",
    },
    {
        "program": "HCS_TXHML",
        "cost_area": "CENTRAL_OFFICE",
        "source_file": "2024-hcs-cost-report-instructions.txt",
        "source_line_start": 4114,
        "source_line_end": 4136,
        "description": "Allocated shared administrative costs reported at actual cost with no markup.",
    },
]

ICF_COST_AREAS = [
    {
        "program": "ICF_IID",
        "cost_area": "RESIDENTIAL_SMALL_MEDIUM_LARGE",
        "source_file": "2024-icf-cost-report-instructions.txt",
        "source_line_start": 3950,
        "source_line_end": 3961,
        "description": "Residential facility costs reported separately by small, medium, and large homes.",
    },
    {
        "program": "ICF_IID",
        "cost_area": "DAY_HABILITATION",
        "source_file": "2024-icf-cost-report-instructions.txt",
        "source_line_start": 3963,
        "source_line_end": 3979,
        "description": "Day habilitation facility and shared-service costs.",
    },
    {
        "program": "ICF_IID",
        "cost_area": "PROGRAM_ADMINISTRATION_OPERATIONS",
        "source_file": "2024-icf-cost-report-instructions.txt",
        "source_line_start": 3981,
        "source_line_end": 3989,
        "description": "Administrative expenses directly chargeable to the facility itself.",
    },
    {
        "program": "ICF_IID",
        "cost_area": "CENTRAL_OFFICE",
        "source_file": "2024-icf-cost-report-instructions.txt",
        "source_line_start": 3991,
        "source_line_end": 4013,
        "description": "Allocated shared administrative costs reported at actual cost with no markup.",
    },
]


def _normalize_program(value: str | None) -> str | None:
    """Map raw program name variants to a canonical key, or return None if unrecognised."""
    if value is None:
        return None
    cleaned = re.sub(r"[^A-Z+]", "", str(value).strip().upper())
    normalized_map = {
        "CLASS": "CLASS",
        "DBMD": "DBMD",
        "HCS": "HCS",
        "MDCP": "MDCP",
        "STAR+": "STAR_PLUS_HCBS",
        "STARPLUS": "STAR_PLUS_HCBS",
        "STAR+PLUSHCBS": "STAR_PLUS_HCBS",
        "STARPLUSHCBS": "STAR_PLUS_HCBS",
        "TXHML": "TXHML",
        "TOTALS": "ALL_PROGRAMS",
    }
    return normalized_map.get(cleaned)


def _clean_int(value: object) -> int | None:
    """Coerce a cell value to int, stripping commas and whitespace. Returns None for blanks."""
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if value == "":
            return None
    return int(float(value))


def _clean_float(value: object) -> float | None:
    """Coerce a cell value to float. Returns None for blanks."""
    if value is None or value == "":
        return None
    return float(value)


def _parse_currency(value: str) -> float:
    """Parse a dollar string like '$1,234.56' into a float."""
    return float(value.replace("$", "").replace(",", ""))


def _interest_list_report_month(path: Path) -> str:
    """Extract a YYYY-MM-01 report month string from an interest-list filename."""
    match = re.search(r"interest-list-data-([a-z]+)-(\d{4})\.xlsx$", path.name)
    if not match:
        raise ValueError(f"Could not parse report month from {path.name}")
    month_name, year = match.groups()
    month = MONTH_MAP[month_name.lower()]
    return f"{year}-{month:02d}-01"


def _stripped_lines(path: Path) -> list[str]:
    """Read a text file and return non-empty lines with whitespace stripped."""
    return [line.strip() for line in path.read_text().splitlines() if line.strip()]


def _find_row(ws, expected: str) -> int:
    """Return the row number whose column-1 value exactly matches *expected*."""
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if value == expected:
            return row
    raise ValueError(f"Row starting with {expected!r} not found in {ws.title}")


def _find_row_prefix(ws, prefix: str) -> int:
    """Return the row number whose column-1 value starts with *prefix*."""
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.strip().startswith(prefix):
            return row
    raise ValueError(f"Row starting with {prefix!r} not found in {ws.title}")


def _find_row_contains_prefix(ws, prefix: str) -> int:
    """Return the row number where any cell's text starts with *prefix*."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.strip().startswith(prefix):
                return row
    raise ValueError(f"Row containing {prefix!r} not found in {ws.title}")


def _find_line_prefix(lines: list[str], prefix: str, start: int = 0) -> int:
    """Return the index of the first line starting with *prefix*, searching from *start*."""
    for idx in range(start, len(lines)):
        if lines[idx].startswith(prefix):
            return idx
    raise ValueError(f"Line starting with {prefix!r} not found")


def _sheet_by_prefix(workbook, prefix: str):
    """Return the first worksheet whose name equals or starts with *prefix*."""
    for sheet_name in workbook.sheetnames:
        if sheet_name == prefix or sheet_name.startswith(prefix):
            return workbook[sheet_name]
    raise KeyError(f"Worksheet starting with {prefix!r} does not exist")


def _program_columns(ws, program_row: int, value_row: int, expected_metric: str) -> list[tuple[int, str]]:
    """Identify (column, program) pairs matching *expected_metric* in the header rows."""
    columns: list[tuple[int, str]] = []
    for col in range(2, ws.max_column + 1):
        program = _normalize_program(ws.cell(program_row, col).value)
        metric = ws.cell(value_row, col).value
        if not program or not isinstance(metric, str):
            continue
        if metric.replace(" ", "").startswith(expected_metric.replace(" ", "")):
            columns.append((col, program))
    return columns


def _row_texts(ws, row: int) -> list[str]:
    """Collect all non-empty, whitespace-normalised text values from a worksheet row."""
    texts: list[str] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.strip():
            texts.append(" ".join(value.split()))
    return texts


def _row_label_and_reference(ws, row: int, label_prefix: str) -> tuple[str | None, str | None]:
    """Extract the label and reference note from a row, keyed by *label_prefix*."""
    texts = _row_texts(ws, row)
    label = next((text for text in texts if text.startswith(label_prefix)), None)
    reference = next(
        (
            text
            for text in texts
            if text != label and _normalize_program(text) is None
        ),
        None,
    )
    return label, reference


def _load_interest_list_workbooks(
    monthly_dir: Path,
) -> list[tuple[Path, object]]:
    """Load all interest-list Excel workbooks from a directory once.

    Returns a sorted list of (path, workbook) tuples.  Callers should
    call ``wb.close()`` on each workbook when finished.
    """
    workbooks = []
    for path in sorted(monthly_dir.glob("interest-list-data-*.xlsx")):
        wb = load_workbook(path, data_only=True)
        workbooks.append((path, wb))
    return workbooks


def build_interest_list_totals(
    monthly_dir: Path,
    *,
    _workbooks: list[tuple[Path, object]] | None = None,
) -> pd.DataFrame:
    rows: list[dict] = []
    workbooks = _workbooks or _load_interest_list_workbooks(monthly_dir)
    for path, wb in workbooks:
        ws = _sheet_by_prefix(wb, "Time on Interest List")
        header_row = _find_row(ws, "Years on List")
        totals_row = _find_row(ws, "Totals")
        if _normalize_program(ws.cell(header_row, 2).value):
            program_row = header_row
            value_row = header_row + 1
        else:
            program_row = header_row - 1
            value_row = header_row
        report_month = _interest_list_report_month(path)
        for col, program in _program_columns(ws, program_row, value_row, "Individuals"):
            rows.append(
                {
                    "report_month": report_month,
                    "program": program,
                    "individuals": _clean_int(ws.cell(totals_row, col).value),
                    "source_file": path.name,
                }
            )
        if _workbooks is None:
            wb.close()
    return pd.DataFrame(rows).sort_values(["report_month", "program"]).reset_index(drop=True)


def build_interest_list_years_on_list(
    monthly_dir: Path,
    *,
    _workbooks: list[tuple[Path, object]] | None = None,
) -> pd.DataFrame:
    rows: list[dict] = []
    workbooks = _workbooks or _load_interest_list_workbooks(monthly_dir)
    for path, wb in workbooks:
        ws = _sheet_by_prefix(wb, "Time on Interest List")
        header_row = _find_row(ws, "Years on List")
        totals_row = _find_row(ws, "Totals")
        if _normalize_program(ws.cell(header_row, 2).value):
            program_row = header_row
            value_row = header_row + 1
            data_start = header_row + 2
        else:
            program_row = header_row - 1
            value_row = header_row
            data_start = header_row + 1
        program_columns = _program_columns(ws, program_row, value_row, "Individuals")
        report_month = _interest_list_report_month(path)
        for row in range(data_start, totals_row):
            bucket = ws.cell(row, 1).value
            if not bucket:
                continue
            for col, program in program_columns:
                individuals = _clean_int(ws.cell(row, col).value)
                percent = _clean_float(ws.cell(row, col + 1).value)
                if individuals is None:
                    continue
                rows.append(
                    {
                        "report_month": report_month,
                        "program": program,
                        "years_on_list_bucket": bucket,
                        "individuals": individuals,
                        "percent_of_program_total": percent,
                        "source_file": path.name,
                    }
                )
        if _workbooks is None:
            wb.close()
    return pd.DataFrame(rows).sort_values(
        ["report_month", "program", "years_on_list_bucket"]
    ).reset_index(drop=True)


def build_interest_list_closure_summary(
    monthly_dir: Path,
    *,
    _workbooks: list[tuple[Path, object]] | None = None,
) -> pd.DataFrame:
    rows: list[dict] = []
    workbooks = _workbooks or _load_interest_list_workbooks(monthly_dir)
    for path, wb in workbooks:
        ws = _sheet_by_prefix(wb, "Closure Overview")
        header_row = _find_row(ws, "Categories:")
        totals_row = _find_row(ws, "Totals")
        if _normalize_program(ws.cell(header_row, 2).value):
            program_row = header_row
            value_row = header_row + 1
            data_start = header_row + 2
        else:
            program_row = header_row - 1
            value_row = header_row
            data_start = header_row + 1
        program_columns = _program_columns(ws, program_row, value_row, "Closures")
        report_month = _interest_list_report_month(path)
        for row in range(data_start, totals_row + 1):
            category = ws.cell(row, 1).value
            for col, program in program_columns:
                closures = _clean_int(ws.cell(row, col).value)
                percent = _clean_float(ws.cell(row, col + 1).value)
                if closures is None:
                    continue
                rows.append(
                    {
                        "report_month": report_month,
                        "program": program,
                        "closure_category": category,
                        "closures": closures,
                        "percent_of_program_closures": percent,
                        "source_file": path.name,
                    }
                )
        if _workbooks is None:
            wb.close()
    return pd.DataFrame(rows).sort_values(
        ["report_month", "program", "closure_category"]
    ).reset_index(drop=True)


def build_interest_list_releases_summary(
    monthly_dir: Path,
    *,
    _workbooks: list[tuple[Path, object]] | None = None,
) -> pd.DataFrame:
    rows: list[dict] = []
    metric_prefixes = [
        ("previous_biennium_counts", "Previous Biennium Counts", False),
        ("enrolled", "Enrolled", False),
        ("denied_declined_withdrawn", "Denied/Declined/Withdrawn", False),
        ("pipeline", "Pipeline", False),
        ("not_categorized", "Not Categorized", True),
        ("total_releases_this_biennium", "Total Releases This Biennium", False),
        ("added_this_biennium", "Added This Biennium", False),
        ("current_interest_list_counts", "Current Interest List Counts", False),
    ]
    workbooks = _workbooks or _load_interest_list_workbooks(monthly_dir)
    for path, wb in workbooks:
        ws = _sheet_by_prefix(wb, "IL Releases Summary")
        header_row = next(
            row
            for row in range(1, ws.max_row + 1)
            if sum(
                1
                for col in range(1, ws.max_column + 1)
                if _normalize_program(ws.cell(row, col).value)
            )
            >= 4
        )
        program_columns = [
            (col, program)
            for col in range(1, ws.max_column + 1)
            if (program := _normalize_program(ws.cell(header_row, col).value))
        ]
        report_month = _interest_list_report_month(path)
        for metric, prefix, allow_missing in metric_prefixes:
            try:
                row = _find_row_contains_prefix(ws, prefix)
            except ValueError:
                if allow_missing:
                    continue
                raise
            label, reference = _row_label_and_reference(ws, row, prefix)
            for col, program in program_columns:
                value = _clean_int(ws.cell(row, col).value)
                if value is None:
                    continue
                rows.append(
                    {
                        "report_month": report_month,
                        "program": program,
                        "metric": metric,
                        "label": label,
                        "reference_text": reference,
                        "value": value,
                        "source_file": path.name,
                    }
                )
        if _workbooks is None:
            wb.close()
    return pd.DataFrame(rows).sort_values(
        ["report_month", "program", "metric"]
    ).reset_index(drop=True)


def build_interest_list_legislative_allocations(page_html: Path) -> pd.DataFrame:
    soup = BeautifulSoup(page_html.read_text(), "html.parser")
    table = soup.find("table")
    if table is None:
        raise ValueError("Legislative allocation table not found on interest-list page")
    rows: list[dict] = []
    for tr in table.find_all("tr")[1:]:
        tds = tr.find_all("td")
        if len(tds) != 2:
            continue
        program_text = " ".join(tds[0].get_text(" ", strip=True).split())
        allocation_text = " ".join(tds[1].get_text(" ", strip=True).split())
        rows.append(
            {
                "biennium": "2022-2023",
                "program_label": program_text,
                "program": _normalize_program(program_text.split("(")[-1].rstrip(")")) or (
                    "ALL_PROGRAMS" if "Total of allocations" in program_text else program_text
                ),
                "legislative_allocation": _clean_int(allocation_text.replace(",", "")),
                "source_file": page_html.name,
            }
        )
    return pd.DataFrame(rows)


def build_setting_costs(report_text: Path) -> pd.DataFrame:
    lines = _stripped_lines(report_text)
    rows: list[dict] = []
    block_labels = [label for label, _, _, _ in SETTING_BLOCKS]
    for block_index, (label, setting, metric_labels, metric_names) in enumerate(SETTING_BLOCKS):
        idx = lines.index(label)
        next_idx = (
            lines.index(block_labels[block_index + 1])
            if block_index + 1 < len(block_labels)
            else _find_line_prefix(lines, "Table 2.", idx + 1)
        )
        block_lines = lines[idx + 1 : next_idx]
        for metric_label in metric_labels:
            if metric_label not in block_lines:
                raise ValueError(f"Expected {metric_label!r} after {label!r}")
        currencies = [line for line in block_lines if line.startswith("$")]
        if len(currencies) != len(metric_names):
            raise ValueError(f"Expected {len(metric_names)} currency values for {label!r}")
        for metric_name, value in zip(metric_names, currencies, strict=True):
            rows.append(
                {
                    "fiscal_year": 2023,
                    "setting": setting,
                    "metric": metric_name,
                    "monthly_average_cost": _parse_currency(value),
                    "source_file": report_text.name,
                }
            )
    return pd.DataFrame(rows)


def build_icf_community_costs_by_lon(report_text: Path) -> pd.DataFrame:
    lines = _stripped_lines(report_text)
    table2_headers = ["Small", "Medium", "Large", "Total"]
    table3_headers = ["Small", "Medium", "Large", "Average"]
    counts: list[dict] = []
    costs: list[dict] = []
    table2_start = _find_line_prefix(
        lines, "Table 2. Non-state ICF/IID by Facility Size: Average Number of Individuals Served Per"
    )
    table3_start = _find_line_prefix(
        lines, "Table 3. Non-state ICF/IID by Facility Size: Monthly Average Cost Per Individual"
    )

    for label, lon in LON_LABELS.items():
        if label == "Overall - Average":
            continue
        idx = _find_line_prefix(lines, label, table2_start)
        values = lines[idx - 4 : idx] if label == "Overall - Total" else lines[idx + 1 : idx + 5]
        for offset, facility_size in enumerate(table2_headers, start=1):
            normalized_size = "ALL" if facility_size == "Total" else facility_size.upper()
            counts.append(
                {
                    "level_of_need": lon,
                    "facility_size": normalized_size,
                    "average_individuals_served_per_month": _clean_int(values[offset - 1].replace(",", "")),
                }
            )

    for label, lon in LON_LABELS.items():
        if label == "Overall - Total":
            continue
        idx = _find_line_prefix(lines, label, table3_start)
        for offset, facility_size in enumerate(table3_headers, start=1):
            normalized_size = "ALL" if facility_size == "Average" else facility_size.upper()
            costs.append(
                {
                    "level_of_need": lon,
                    "facility_size": normalized_size,
                    "monthly_average_cost_per_individual": _parse_currency(lines[idx + offset]),
                }
            )

    return pd.merge(pd.DataFrame(counts), pd.DataFrame(costs), on=["level_of_need", "facility_size"])


def build_residential_rate_components(rate_csv: Path) -> pd.DataFrame:
    df = pd.read_csv(rate_csv)

    hcs = df[(df["program"] == "HCS") & (df["service_category"] == "SL/RSS")].copy()
    hcs["program"] = "HCS"
    hcs["setting"] = "SL_RSS"
    hcs["facility_size"] = pd.NA
    hcs["level_of_need"] = hcs["bill_description"].str.extract(r"LON (\d+)").astype(int)

    icf = df[
        (df["program"] == "ICF")
        & (df["bill_description"].str.contains("NON-STATE COMMUNITY RESIDENTIAL", na=False))
    ].copy()
    icf["program"] = "ICF_IID"
    icf["setting"] = "NON_STATE_COMMUNITY_RESIDENTIAL"
    icf["facility_size"] = icf["bill_description"].str.extract(r"- (Small|Medium|Large)$")[0].str.upper()
    icf["level_of_need"] = icf["bill_description"].str.extract(r"LON (\d+)").astype(int)

    combined = pd.concat([hcs, icf], ignore_index=True, sort=False)
    combined["non_attendant_reimbursement_component"] = (
        combined["current_rate"] - combined["attendant_cost"]
    )
    return combined[
        [
            "program",
            "setting",
            "facility_size",
            "level_of_need",
            "bill_description",
            "current_rate",
            "attendant_cost",
            "attendant_cost_pct",
            "non_attendant_reimbursement_component",
            "wage_required",
            "rate_effective_date",
        ]
    ].sort_values(["program", "facility_size", "level_of_need"]).reset_index(drop=True)


def build_cost_report_cost_areas() -> pd.DataFrame:
    return pd.DataFrame(HCS_COST_AREAS + ICF_COST_AREAS)


# ---------------------------------------------------------------------------
# Sept 2025 rate change comparison
# Source: 9-1-2025-payment-rate-actions.pdf, Tables 1 & 2
# GAA SB1, 89th Legislature, Rider 23: $13.00/hr wage, 15% PTB facility,
# 14% PTB non-facility, +$0.24/hr admin increase
# ---------------------------------------------------------------------------

# Policy parameters from Rider 23 (page 3 of the PDF)
RIDER_23_WAGE = 13.00
RIDER_23_PTB_FACILITY = 0.15
RIDER_23_PTB_NON_FACILITY = 0.14
RIDER_23_ADMIN_INCREASE_PER_HOUR = 0.24
RATE_EFFECTIVE_DATE_NEW = "2025-09-01"

# Table 1: HCS SL/RSS proposed rates (page 6)
_HCS_SL_RSS_PROPOSED = [
    # (lon, cur_att, cur_adm, cur_total, new_att, new_adm, new_total, pct)
    (1, 78.04, 62.96, 149.08, 91.14, 64.44, 163.66, 0.10),
    (5, 85.74, 63.23, 158.00, 101.81, 64.88, 175.72, 0.11),
    (8, 97.03, 63.61, 171.04, 117.37, 65.51, 193.28, 0.13),
    (6, 115.88, 64.22, 192.80, 143.31, 66.54, 222.55, 0.15),
    (9, 191.63, 66.71, 280.24, 247.05, 70.71, 339.66, 0.21),
]

# Table 2: ICF/IID proposed per-diem rates (page 8)
# (size, lon, dh_dc, dh_other, dh_admin, res_dc, res_other, res_admin, total_daily, pct_increase)
_ICF_IID_PROPOSED = [
    ("SMALL", 1, 16.56, 1.22, 5.72, 61.33, 22.44, 66.34, 173.61, 0.12),
    ("SMALL", 5, 18.57, 1.53, 7.18, 69.22, 22.34, 73.04, 191.88, 0.11),
    ("SMALL", 8, 21.89, 2.03, 9.56, 86.11, 22.21, 78.92, 220.72, 0.12),
    ("SMALL", 6, 28.64, 3.05, 14.33, 120.35, 27.58, 82.40, 276.35, 0.14),
    ("SMALL", 9, 77.89, 10.08, 57.30, 244.43, 26.26, 95.87, 511.83, 0.25),
    ("MEDIUM", 1, 16.56, 1.22, 5.72, 59.20, 19.68, 52.80, 155.18, 0.22),
    ("MEDIUM", 5, 18.57, 1.53, 7.18, 69.22, 19.80, 59.03, 175.33, 0.22),
    ("MEDIUM", 8, 21.89, 2.03, 9.56, 86.11, 19.94, 67.68, 207.21, 0.22),
    ("MEDIUM", 6, 28.64, 3.05, 14.33, 120.35, 29.59, 68.96, 264.92, 0.30),
    ("MEDIUM", 9, 77.89, 10.08, 57.30, 244.43, 24.51, 79.31, 493.52, 0.26),
    ("LARGE", 1, 16.56, 1.22, 5.72, 59.20, 15.56, 60.90, 159.16, 0.28),
    ("LARGE", 5, 18.57, 1.53, 7.18, 69.22, 16.12, 61.65, 174.27, 0.31),
    ("LARGE", 8, 21.89, 2.03, 9.56, 86.11, 16.89, 64.87, 201.35, 0.36),
    ("LARGE", 6, 28.64, 3.05, 14.33, 120.35, 33.42, 66.49, 266.28, 0.33),
    ("LARGE", 9, 77.89, 9.99, 57.30, 244.43, 44.48, 73.12, 507.21, 0.29),
]


def build_proposed_rate_comparison(old_rates_csv: Path) -> pd.DataFrame:
    """Build old-vs-new rate comparison for HCS SL/RSS and ICF/IID.

    Joins the existing 2023-09-01 rates from the wage calculator with the
    proposed 2025-09-01 rates from the payment rate actions PDF.

    Source: 9-1-2025-payment-rate-actions.pdf, Tables 1 & 2.
    """
    old = pd.read_csv(old_rates_csv)

    rows = []

    # --- HCS SL/RSS (Table 1) ---
    for lon, cur_att, cur_adm, cur_total, new_att, new_adm, new_total, pct in _HCS_SL_RSS_PROPOSED:
        # Get the old rate from checked-in data for cross-check
        old_match = old[
            (old["program"] == "HCS")
            & (old["setting"] == "SL_RSS")
            & (old["level_of_need"] == lon)
        ]
        old_rate = float(old_match["current_rate"].iloc[0]) if len(old_match) > 0 else cur_total

        for svc in ["SL", "RSS"]:
            rows.append({
                "program": "HCS",
                "setting": "SL_RSS",
                "facility_size": pd.NA,
                "level_of_need": lon,
                "service_type": svc,
                "old_rate": old_rate,
                "old_attendant_cost": cur_att,
                "old_admin_cost": cur_adm,
                "new_rate": new_total,
                "new_attendant_cost": new_att,
                "new_admin_cost": new_adm,
                "rate_increase_dollar": new_total - cur_total,
                "rate_increase_pct": pct,
                "new_attendant_cost_pct": round(new_att / new_total, 4),
                "old_rate_effective_date": "2023-09-01",
                "new_rate_effective_date": RATE_EFFECTIVE_DATE_NEW,
            })

    # --- ICF/IID (Table 2) ---
    for row in _ICF_IID_PROPOSED:
        size, lon, dh_dc, dh_other, dh_adm, res_dc, res_other, res_adm, total, pct = row
        old_match = old[
            (old["program"] == "ICF_IID")
            & (old["facility_size"] == size)
            & (old["level_of_need"] == lon)
        ]
        old_rate = float(old_match["current_rate"].iloc[0]) if len(old_match) > 0 else None
        # Attendant cost in ICF = day hab direct care + residential direct care
        new_attendant = dh_dc + res_dc
        old_att = float(old_match["attendant_cost"].iloc[0]) if len(old_match) > 0 else None

        rows.append({
            "program": "ICF_IID",
            "setting": "NON_STATE_COMMUNITY_RESIDENTIAL",
            "facility_size": size,
            "level_of_need": lon,
            "service_type": "ICF_PER_DIEM",
            "old_rate": old_rate,
            "old_attendant_cost": old_att,
            "old_admin_cost": float(old_rate - old_att) if old_rate and old_att else None,
            "new_rate": total,
            "new_attendant_cost": new_attendant,
            "new_admin_cost": total - new_attendant,
            "rate_increase_dollar": round(total - old_rate, 2) if old_rate else None,
            "rate_increase_pct": pct,
            "new_attendant_cost_pct": round(new_attendant / total, 4),
            "old_rate_effective_date": "2023-09-01",
            "new_rate_effective_date": RATE_EFFECTIVE_DATE_NEW,
        })

    df = pd.DataFrame(rows)
    return df.sort_values(
        ["program", "facility_size", "level_of_need", "service_type"]
    ).reset_index(drop=True)


# ---------------------------------------------------------------------------
# ASPE state-level DCW wage data (December 2024 issue brief)
# Source: aspe-dcw-wages-brief.pdf, Tables A1 & A2
# Data: BLS OEWS May 2023, all wages in 2023 dollars
# ---------------------------------------------------------------------------

# Table A1: HHA/PCA 2023 wages — (state, median_wage, avg_entry_level, gap)
_ASPE_HHA_PCA_2023 = [
    ("Alabama", 11.45, 16.58, 5.13),
    ("Alaska", 17.81, 21.54, 3.73),
    ("Arizona", 16.01, 19.34, 3.33),
    ("Arkansas", 13.20, 16.78, 3.58),
    ("California", 16.12, 21.01, 4.89),
    ("Colorado", 17.37, 20.98, 3.61),
    ("Connecticut", 17.63, 20.10, 2.47),
    ("Delaware", 14.18, 18.67, 4.49),
    ("District of Columbia", 17.79, 22.24, 4.45),
    ("Florida", 14.92, 17.51, 2.59),
    ("Georgia", 13.43, 17.27, 3.84),
    ("Hawaii", 17.26, 20.81, 3.55),
    ("Idaho", 14.24, 18.09, 3.85),
    ("Illinois", 16.84, 19.78, 2.94),
    ("Indiana", 14.44, 18.68, 4.24),
    ("Iowa", 16.52, 18.67, 2.15),
    ("Kansas", 13.43, 18.09, 4.66),
    ("Kentucky", 14.60, 17.49, 2.89),
    ("Louisiana", 10.00, 16.61, 6.61),
    ("Maine", 17.39, 19.64, 2.25),
    ("Maryland", 16.72, 19.59, 2.87),
    ("Massachusetts", 18.00, 21.34, 3.34),
    ("Michigan", 14.84, 18.68, 3.84),
    ("Minnesota", 16.56, 20.50, 3.94),
    ("Mississippi", 10.97, 15.67, 4.70),
    ("Missouri", 13.80, 18.24, 4.44),
    ("Montana", 14.78, 18.19, 3.41),
    ("Nebraska", 15.31, 18.49, 3.18),
    ("Nevada", 13.34, 18.35, 5.01),
    ("New Hampshire", 16.55, 19.74, 3.19),
    ("New Jersey", 17.07, 20.67, 3.60),
    ("New Mexico", 12.83, 17.42, 4.59),
    ("New York", 17.52, 21.38, 3.86),
    ("North Carolina", 13.81, 17.19, 3.38),
    ("North Dakota", 17.89, 20.64, 2.75),
    ("Ohio", 14.08, 18.36, 4.28),
    ("Oklahoma", 12.33, 16.55, 4.22),
    ("Oregon", 18.50, 20.32, 1.82),
    ("Pennsylvania", 13.94, 18.78, 4.84),
    ("Rhode Island", 18.25, 19.43, 1.18),
    ("South Carolina", 13.62, 16.94, 3.32),
    ("South Dakota", 16.68, 17.66, 0.98),
    ("Tennessee", 13.77, 17.57, 3.80),
    ("Texas", 10.83, 17.60, 6.77),
    ("Utah", 16.84, 18.85, 2.01),
    ("Vermont", 15.54, 20.51, 4.97),
    ("Virginia", 13.01, 18.55, 5.54),
    ("Washington", 20.35, 22.52, 2.17),
    ("West Virginia", 12.34, 16.53, 4.19),
    ("Wisconsin", 15.30, 19.22, 3.92),
    ("Wyoming", 14.33, 19.22, 4.89),
]


def build_aspe_state_wages() -> pd.DataFrame:
    """Build state-level HHA/PCA wage comparison from ASPE brief.

    Source: aspe-dcw-wages-brief.pdf, Table A1 (pages 11-13).
    All wages in 2023 dollars. BLS OEWS May 2023.
    """
    rows = []
    for state, median, entry_avg, gap in _ASPE_HHA_PCA_2023:
        rows.append({
            "state": state,
            "occupation": "HHA_PCA",
            "median_wage": median,
            "avg_entry_level_wage": entry_avg,
            "wage_gap": gap,
            "wage_ratio": round(median / entry_avg, 4),
            "data_year": 2023,
            "source": "ASPE Issue Brief Dec 2024, Table A1",
        })
    return pd.DataFrame(rows).sort_values("state").reset_index(drop=True)


# ---------------------------------------------------------------------------
# PHI national DCW wage trends (Key Facts 2025)
# Source: phi-dcw-key-facts-2025.pdf, Executive Summary (pages 3-4)
# Wages adjusted for inflation using 2024 dollars
# ---------------------------------------------------------------------------

# (setting, wage_2014, wage_2024)
_PHI_DCW_WAGES = [
    ("home_care", 13.07, 16.77),
    ("residential_care", 14.10, 17.71),
    ("nursing_homes", 15.37, 18.83),
    ("all_dcw", 14.38, 17.36),
]

# Workforce size (2024)
PHI_HOME_CARE_WORKERS = 3_200_000
PHI_RESIDENTIAL_CARE_AIDES = 687_000
PHI_NURSING_ASSISTANTS = 492_000


def build_phi_national_dcw_trends() -> pd.DataFrame:
    """Build national DCW wage trend data from PHI Key Facts 2025.

    Source: phi-dcw-key-facts-2025.pdf, Executive Summary.
    Wages in inflation-adjusted 2024 dollars.
    """
    rows = []
    for setting, w2014, w2024 in _PHI_DCW_WAGES:
        for year, wage in [(2014, w2014), (2024, w2024)]:
            rows.append({
                "setting": setting,
                "year": year,
                "median_wage": wage,
                "source": "PHI Key Facts 2025",
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Rate Enhancement (ACRE/DCSE) Evaluation (October 2024)
# Source: rate-enhancement-evaluation-2024.pdf (Rider 30(d) report)
# ---------------------------------------------------------------------------

# Table 1: Program size by provider type (SFY)
# (provider_type, 2019, 2020, 2021, 2022, 2023, 2024)
_ACRE_PROGRAM_SIZE = [
    ("Community Care", 111_886_958, 133_213_666, 135_193_874,
     129_986_526, 157_641_842, 163_285_145),
    ("IDD Providers", 11_922_770, 13_738_973, 48_050_497,
     52_512_594, 48_089_033, 67_623_973),
    ("Nursing Facility", 74_981_201, 91_504_971, 65_036_165,
     56_227_146, 42_390_182, 75_350_493),
]

# Table 2: DCSE Nursing Facility participation
# (participating, total, pct_participating, pct_utilization)
ACRE_DCSE_NURSING = (1107, 1187, 0.93, 0.9020)

# Table 3: Attendant Comp participation (SFY 2024)
# (program, participating, total, pct, pct_utilization)
_ACRE_PARTICIPATION = [
    ("PHC_CAS_FC", 1840, 2096, 0.88, 0.9611),
    ("DAHS", 252, 282, 0.89, 0.9043),
    ("RC", 20, 52, 0.38, 0.3472),
    ("CLASS", 103, 116, 0.89, 0.6642),
    ("DBMD", 37, 49, 0.76, 0.8138),
    ("ICF_IID", 88, 125, 0.70, 0.6806),
    ("HCS_TxHmL", 241, 696, 0.35, 0.5294),
]

# Efficacy: wage differential participating vs non-participating
# (provider_type, diff_pct_low, diff_pct_high, note)
_ACRE_WAGE_DIFFERENTIAL = [
    ("DAHS_CLASS_PHC_CAS_FC", 0.01, 0.04,
     "Negligible; high participation = small control group"),
    ("ICF_IID", 0.07, 0.07, "7% differential"),
    ("HCS_TxHmL", 0.16, 0.16,
     "16% differential; low participation"),
]

# Table 4: 2019-20 Recoupment by report type
# (type, entities, recouped, pct, sum_dollars)
_ACRE_RECOUPMENTS = [
    ("CPC (CLASS/CAS/FC/PHC)", 2099, 89, 0.042, 4_050_086),
    ("DAHS", 490, 43, 0.088, 94_486),
    ("DBMD", 19, 7, 0.36, 141_113),
    ("HCS/TxHmL", 712, 55, 0.077, 1_066_166),
    ("ICF/IID", 130, 25, 0.192, 1_219_442),
    ("Nursing Facility", 1172, 138, 0.12, 8_323_355),
]

# Table 5: Attendant Base Wage History
# (effective_date, fed_min, att_base, citation)
_ACRE_WAGE_HISTORY = [
    ("2000-09-01", 5.15, 5.15, None),
    ("2007-07-24", 5.85, 5.15, None),
    ("2007-08-01", 5.85, 5.85, None),
    ("2008-07-24", 6.55, 5.85, None),
    ("2008-08-01", 6.55, 6.55, None),
    ("2009-07-24", 7.25, 6.55, None),
    ("2009-08-01", 7.25, 7.25, None),
    ("2013-09-01", 7.25, 7.50,
     "S.B. 1, 83rd Leg, 2013, Art II Sec 61"),
    ("2014-09-01", 7.25, 7.86,
     "S.B. 1, 83rd Leg, 2013, Art II Sec 61"),
    ("2015-09-01", 7.25, 8.00,
     "H.B. 1, 84th Leg, 2015, Art II Sec 47"),
    ("2019-09-01", 7.25, 8.11,
     "H.B. 1, 86th Leg, 2019, Art II Rider 45"),
    ("2023-09-01", 7.25, 10.60,
     "S.B. 1, 88th Leg, 2023, Art II Rider 30(a)"),
]


def build_acre_evaluation() -> pd.DataFrame:
    """Build ACRE/DCSE participation and efficacy summary.

    Source: rate-enhancement-evaluation-2024.pdf, Tables 2-3.
    """
    rows = []
    part, total, pct_p, pct_u = ACRE_DCSE_NURSING
    rows.append({
        "program": "Nursing Facility (DCSE)",
        "enhancement_type": "Direct Care Staff Enhancement",
        "participating_providers": part,
        "total_contracted_providers": total,
        "pct_participating": pct_p,
        "pct_utilization": pct_u,
        "wage_diff_pct_low": None,
        "wage_diff_pct_high": None,
    })
    for prog, part, total, pct_p, pct_u in _ACRE_PARTICIPATION:
        diff_low = diff_high = None
        for wt, lo, hi, _ in _ACRE_WAGE_DIFFERENTIAL:
            if prog.startswith("ICF") and wt == "ICF_IID":
                diff_low, diff_high = lo, hi
            elif prog.startswith("HCS") and wt == "HCS_TxHmL":
                diff_low, diff_high = lo, hi
            elif wt == "DAHS_CLASS_PHC_CAS_FC" and prog in (
                "PHC_CAS_FC", "DAHS", "CLASS",
            ):
                diff_low, diff_high = lo, hi
        rows.append({
            "program": prog,
            "enhancement_type": (
                "Attendant Compensation Rate Enhancement"
            ),
            "participating_providers": part,
            "total_contracted_providers": total,
            "pct_participating": pct_p,
            "pct_utilization": pct_u,
            "wage_diff_pct_low": diff_low,
            "wage_diff_pct_high": diff_high,
        })
    return pd.DataFrame(rows)


def build_acre_program_funding() -> pd.DataFrame:
    """Build ACRE program funding by provider type (SFY 2019-2024).

    Source: rate-enhancement-evaluation-2024.pdf, Table 1.
    """
    rows = []
    years = [2019, 2020, 2021, 2022, 2023, 2024]
    for provider_type, *amounts in _ACRE_PROGRAM_SIZE:
        for year, amount in zip(years, amounts):
            rows.append({
                "provider_type": provider_type,
                "state_fiscal_year": year,
                "estimated_program_size": amount,
                "source": (
                    "HHSC Rate Enhancement Evaluation "
                    "Oct 2024, Table 1"
                ),
            })
    return pd.DataFrame(rows)


def build_acre_recoupments() -> pd.DataFrame:
    """Build ACRE recoupment data (2019-20 cost reports).

    Source: rate-enhancement-evaluation-2024.pdf, Table 4.
    """
    rows = []
    for rtype, entities, recouped, pct, total in _ACRE_RECOUPMENTS:
        rows.append({
            "report_type": rtype,
            "count_entities": entities,
            "count_recouped": recouped,
            "pct_recouped": pct,
            "sum_recoupments": total,
            "source": (
                "HHSC Rate Enhancement Evaluation "
                "Oct 2024, Table 4"
            ),
        })
    return pd.DataFrame(rows)


def build_acre_wage_history() -> pd.DataFrame:
    """Build attendant base wage history vs federal minimum.

    Source: rate-enhancement-evaluation-2024.pdf, Table 5.
    """
    rows = []
    for eff, fed, att, cite in _ACRE_WAGE_HISTORY:
        rows.append({
            "effective_date": eff,
            "federal_minimum_wage": fed,
            "attendant_base_wage": att,
            "legislative_citation": cite,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Federal Funding Context (SFY 2024)
# Source: annual-federal-funds-report-2024.pdf, Figure 1 (p.1)
# ---------------------------------------------------------------------------

# (agency, federal_funds, all_funds, pct_federal)
_FEDERAL_FUNDS_SFY2024 = [
    ("DSHS", 990_888_584, 1_653_610_703, 0.60),
    ("HHSC", 26_869_487_588, 47_850_324_419, 0.56),
]

# Top federal funding sources (p.1 text)
MEDICAID_FEDERAL_AMOUNT = 25_500_000_000  # $25.5B
MEDICAID_FEDERAL_PCT = 0.93  # 93% of all federal HHS funds
CHIP_FEDERAL_AMOUNT = 981_300_000  # $981.3M
CHIP_FEDERAL_PCT = 0.04  # 4% of federal funds


def build_federal_funds_summary() -> pd.DataFrame:
    """Build federal funds summary for TX HHS system.

    Source: annual-federal-funds-report-2024.pdf, Figure 1.
    """
    rows = []
    for agency, fed, total, pct in _FEDERAL_FUNDS_SFY2024:
        rows.append({
            "agency": agency,
            "federal_funds": fed,
            "all_funds": total,
            "pct_federal": pct,
            "state_fiscal_year": 2024,
            "source": (
                "HHSC Annual Federal Funds Report "
                "Dec 2024, Figure 1"
            ),
        })
    # Add total row — use the PDF's adjusted TOTAL (Figure 1, TOTAL row),
    # NOT the sum of agency rows.  The PDF's footnote 2 explains the
    # difference: "Excludes employee benefits, certain payments made as a
    # result of local funding sources (Intergovernmental Transfers), and the
    # value of SNAP benefits."
    rows.append({
        "agency": "TOTAL",
        "federal_funds": 27_426_247_798,
        "all_funds": 49_503_935_122,
        "pct_federal": 0.55,
        "state_fiscal_year": 2024,
        "source": (
            "HHSC Annual Federal Funds Report "
            "Dec 2024, Figure 1 (adjusted TOTAL per footnote 2)"
        ),
    })
    return pd.DataFrame(rows)
