"""Tests verifying the headline statistics cited in the README and policy brief.

Each test locks down a key claim to its source data, so any data update
that changes a headline number is caught immediately.
"""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from texas_hhcs.cpi import deflate_wage

# -- Constants from the investigation --
ROOT = Path(__file__).resolve().parent.parent
RAW_WAGE_CALCULATOR = (
    ROOT / "data" / "raw" / "rates" / "ltss-personal-attendant-base-wage-calculator.xlsx"
)
PROCESSED_BLS = ROOT / "data" / "processed" / "bls_oews_texas_2024.csv"
LBB_RIDER_TEXT = ROOT / "data" / "raw" / "external" / "lbb-88th-article-ii-rider.txt"
NCI_WORKFORCE_TEXT = ROOT / "data" / "raw" / "external" / "nci-state-of-the-workforce-2023.txt"
HHSC_WAGE = 10.60
TX_HHA_EMPLOYMENT = 314_610


@pytest.fixture()
def cpi_annual():
    """CPI-U South annual averages used in the analysis.

    Source: BLS series CUUR0300SA0, annual means for 2015 and 2025.
    Hardcoded so tests run offline without hitting the BLS API.
    """
    return pd.DataFrame({
        "year": [2015, 2025],
        "annual_cpi": [234.812, 318.451],
    })


class TestInflationErosion:
    """$10.60 frozen since ~2015 should be worth ~$14.38 in 2025 dollars,
    and its real purchasing power should be ~$7.82."""

    def test_should_be_wage_approximately_14_38(self, cpi_annual):
        should_be = deflate_wage(HHSC_WAGE, 2015, 2025, cpi_annual)
        assert 14.00 < should_be < 14.80, f"Expected ~$14.38, got ${should_be:.2f}"

    def test_real_purchasing_power_approximately_7_82(self, cpi_annual):
        cpi_2015 = cpi_annual.loc[cpi_annual["year"] == 2015, "annual_cpi"].iloc[0]
        cpi_2025 = cpi_annual.loc[cpi_annual["year"] == 2025, "annual_cpi"].iloc[0]
        real_value = HHSC_WAGE * (cpi_2015 / cpi_2025)
        assert 7.50 < real_value < 8.10, f"Expected ~$7.82, got ${real_value:.2f}"


class TestSourceData:
    """Source-backed figures should come from checked-in data files."""

    def test_hhsc_target_wage_cell_b7_is_10_60(self):
        wb = load_workbook(RAW_WAGE_CALCULATOR, data_only=True, read_only=True)
        ws = wb["Fiscal by Program"]
        assert ws["B7"].value == pytest.approx(HHSC_WAGE)

    def test_bls_soc_31_1120_matches_processed_extract(self):
        df_bls = pd.read_csv(PROCESSED_BLS)
        soc_311120 = df_bls[df_bls["soc_code"] == "31-1120"]
        assert set(soc_311120["occupation"].unique()) == {"Home Health and Personal Care Aides"}

        measures = soc_311120.set_index("measure")["value"]
        assert measures["hourly_mean"] == pytest.approx(12.19)
        assert int(measures["employment"]) == TX_HHA_EMPLOYMENT

    def test_lbb_3292_figure_is_for_txhml_not_hcs(self):
        lbb_text = LBB_RIDER_TEXT.read_text()
        assert "Strategy A.3.4, Texas Home Living" in lbb_text
        assert "fiscal year 2025 for 3,292 end-of-year waiver slots." in lbb_text

    def test_nci_weighted_average_turnover_ratio_is_39_7(self):
        nci_text = NCI_WORKFORCE_TEXT.read_text()
        assert "weighted average turnover ratio was 39.7%." in nci_text


class TestRider23RateChange:
    """The Sept 2025 rate actions set $13.00/hr wage with 15%/14% PTB."""

    def test_new_wage_assumption_is_13(self):
        from texas_hhcs.verified_datasets import RIDER_23_WAGE

        assert RIDER_23_WAGE == pytest.approx(13.00)

    def test_ptb_facility_is_15_pct(self):
        from texas_hhcs.verified_datasets import RIDER_23_PTB_FACILITY

        assert RIDER_23_PTB_FACILITY == pytest.approx(0.15)

    def test_ptb_non_facility_is_14_pct(self):
        from texas_hhcs.verified_datasets import RIDER_23_PTB_NON_FACILITY

        assert RIDER_23_PTB_NON_FACILITY == pytest.approx(0.14)

    def test_hcs_sl_lon1_proposed_rate_is_163_66(self):
        from texas_hhcs.verified_datasets import _HCS_SL_RSS_PROPOSED

        lon1 = [r for r in _HCS_SL_RSS_PROPOSED if r[0] == 1][0]
        assert lon1[6] == pytest.approx(163.66)  # total proposed rate

    def test_icf_small_lon1_proposed_rate_is_173_61(self):
        from texas_hhcs.verified_datasets import _ICF_IID_PROPOSED

        small_lon1 = [r for r in _ICF_IID_PROPOSED if r[0] == "SMALL" and r[1] == 1][0]
        assert small_lon1[8] == pytest.approx(173.61)  # total daily rate

    def test_new_wage_still_below_cpi_parity(self, cpi_annual):
        """$13.00 is still less than the inflation-adjusted $10.60."""
        should_be = deflate_wage(HHSC_WAGE, 2015, 2025, cpi_annual)
        assert 13.00 < should_be, (
            f"New wage $13.00 should still be below CPI parity ${should_be:.2f}"
        )

    def test_wage_increase_is_22_6_pct(self):
        increase_pct = (13.00 - HHSC_WAGE) / HHSC_WAGE * 100
        assert increase_pct == pytest.approx(22.6, abs=0.1)


class TestNationalWageComparison:
    """TX has the largest HHA/PCA wage gap nationally per ASPE."""

    def test_tx_hha_pca_median_is_10_83(self):
        from texas_hhcs.verified_datasets import _ASPE_HHA_PCA_2023

        tx = [r for r in _ASPE_HHA_PCA_2023 if r[0] == "Texas"][0]
        assert tx[1] == pytest.approx(10.83)

    def test_tx_wage_gap_is_6_77_largest_nationally(self):
        from texas_hhcs.verified_datasets import _ASPE_HHA_PCA_2023

        tx = [r for r in _ASPE_HHA_PCA_2023 if r[0] == "Texas"][0]
        assert tx[3] == pytest.approx(6.77)
        # Verify TX has the largest gap
        max_gap = max(r[3] for r in _ASPE_HHA_PCA_2023)
        assert tx[3] == pytest.approx(max_gap)

    def test_phi_national_home_care_median_is_16_77(self):
        from texas_hhcs.verified_datasets import _PHI_DCW_WAGES

        home_care = [r for r in _PHI_DCW_WAGES if r[0] == "home_care"][0]
        assert home_care[2] == pytest.approx(16.77)  # 2024 wage

    def test_phi_all_dcw_median_is_17_36(self):
        from texas_hhcs.verified_datasets import _PHI_DCW_WAGES

        all_dcw = [r for r in _PHI_DCW_WAGES if r[0] == "all_dcw"][0]
        assert all_dcw[2] == pytest.approx(17.36)


class TestRateEnhancementEvaluation:
    """ACRE evaluation figures from Rider 30(d) report Oct 2024."""

    def test_nursing_facility_participation_is_93_pct(self):
        from texas_hhcs.verified_datasets import ACRE_DCSE_NURSING

        assert ACRE_DCSE_NURSING[2] == pytest.approx(0.93)

    def test_hcs_txhml_participation_is_35_pct(self):
        from texas_hhcs.verified_datasets import _ACRE_PARTICIPATION

        hcs = [r for r in _ACRE_PARTICIPATION if r[0] == "HCS_TxHmL"][0]
        assert hcs[3] == pytest.approx(0.35)

    def test_hcs_txhml_wage_differential_is_16_pct(self):
        from texas_hhcs.verified_datasets import _ACRE_WAGE_DIFFERENTIAL

        hcs = [r for r in _ACRE_WAGE_DIFFERENTIAL if r[0] == "HCS_TxHmL"][0]
        assert hcs[2] == pytest.approx(0.16)

    def test_icf_iid_participation_is_70_pct(self):
        from texas_hhcs.verified_datasets import _ACRE_PARTICIPATION

        icf = [r for r in _ACRE_PARTICIPATION if r[0] == "ICF_IID"][0]
        assert icf[3] == pytest.approx(0.70)

    def test_total_recoupments_2019_20(self):
        from texas_hhcs.verified_datasets import _ACRE_RECOUPMENTS

        total = sum(r[4] for r in _ACRE_RECOUPMENTS)
        assert total == pytest.approx(14_894_648)

    def test_attendant_base_wage_2023_is_10_60(self):
        from texas_hhcs.verified_datasets import _ACRE_WAGE_HISTORY

        latest = _ACRE_WAGE_HISTORY[-1]
        assert latest[2] == pytest.approx(10.60)
        assert latest[0] == "2023-09-01"


class TestFederalFundingContext:
    """Federal funding figures from Annual Federal Funds Report."""

    def test_medicaid_is_93_pct_of_federal_hhs(self):
        from texas_hhcs.verified_datasets import MEDICAID_FEDERAL_PCT

        assert MEDICAID_FEDERAL_PCT == pytest.approx(0.93)

    def test_medicaid_federal_amount_is_25_5b(self):
        from texas_hhcs.verified_datasets import MEDICAID_FEDERAL_AMOUNT

        assert MEDICAID_FEDERAL_AMOUNT == pytest.approx(25_500_000_000)

    def test_hhsc_federal_funds_is_26_87b(self):
        from texas_hhcs.verified_datasets import _FEDERAL_FUNDS_SFY2024

        hhsc = [r for r in _FEDERAL_FUNDS_SFY2024 if r[0] == "HHSC"][0]
        assert hhsc[1] == pytest.approx(26_869_487_588)

    def test_hhs_system_is_55_pct_federal(self):
        """Source: annual-federal-funds-report-2024.pdf, Figure 1, TOTAL row.

        The PDF's adjusted TOTAL (footnote 2: excludes employee benefits,
        IGTs, SNAP) shows 55% federal, NOT the 56% from summing agency rows.
        """
        from texas_hhcs.verified_datasets import build_federal_funds_summary

        df = build_federal_funds_summary()
        total_row = df[df["agency"] == "TOTAL"].iloc[0]
        assert total_row["pct_federal"] == pytest.approx(0.55, abs=0.01)
